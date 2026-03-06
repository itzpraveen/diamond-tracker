from datetime import datetime, timezone
from io import BytesIO
import uuid

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload

from app.db import get_db
from app.deps import require_roles
from app.models import Batch, BatchItem, BatchStatus, Branch, Factory, ItemJob, Role, Status, StatusEvent
from app.schemas import BatchAddItem, BatchCreate, BatchDetail, BatchDispatchRequest, BatchOut, JobOut
from app.utils.pdf import generate_manifest_pdf
from app.utils.roles import select_role_for_action
from app.utils.transitions import STATUS_HOLDER_ROLE
from app.utils.vouchers import format_voucher_code, next_voucher_sequence

router = APIRouter(prefix="/batches", tags=["batches"])

MANIFEST_ITEM_COLUMNS = [
    "Job ID",
    "Item Details",
    "Total Weight",
    "Total Carat",
    "Total Value",
    "Voucher No",
    "Customer Name",
    "Phone",
    "Style Number",
    "Card Weight",
    "Factory",
    "Status",
    "Work Narration",
    "Item Source",
    "Repair Type",
    "Target Return Date",
    "Created At",
    "Added To Voucher At",
]


def _get_default_branch(db: Session) -> Branch:
    branch = db.query(Branch).first()
    if not branch:
        branch = Branch(name="Main Branch")
        db.add(branch)
        db.commit()
        db.refresh(branch)
    return branch


def _get_batch(db: Session, batch_id: str, *, with_items: bool = False, with_factory: bool = False) -> Batch:
    try:
        batch_uuid = uuid.UUID(batch_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid voucher id") from exc
    query = db.query(Batch)
    if with_factory:
        query = query.options(selectinload(Batch.factory))
    if with_items:
        query = query.options(
            selectinload(Batch.items)
            .selectinload(BatchItem.job)
            .selectinload(ItemJob.factory)
        )
    batch = query.filter(Batch.id == batch_uuid).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Voucher not found")
    return batch


def _get_job_by_code(db: Session, job_code: str) -> ItemJob:
    job = db.query(ItemJob).filter(ItemJob.job_id == job_code).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


def _get_factory_by_uuid(db: Session, factory_id: uuid.UUID) -> Factory:
    factory = db.query(Factory).filter(Factory.id == factory_id).first()
    if not factory:
        raise HTTPException(status_code=404, detail="Factory not found")
    if not factory.is_active:
        raise HTTPException(status_code=400, detail="Factory is inactive")
    return factory


def _sync_batch_item_count(db: Session, batch: Batch) -> int:
    item_count = (
        db.query(func.count(BatchItem.id))
        .filter(BatchItem.batch_id == batch.id)
        .scalar()
    ) or 0
    batch.item_count = int(item_count)
    if batch.item_count == 0:
        batch.status = BatchStatus.CREATED
        batch.factory_id = None
        batch.dispatch_date = None
        batch.expected_return_date = None
    return batch.item_count


def _remove_batch_item(batch: Batch, batch_item: BatchItem, user) -> None:
    job = batch_item.job
    if job.current_status != Status.DISPATCHED_TO_FACTORY:
        raise HTTPException(
            status_code=400,
            detail=f"Job {job.job_id} can only be removed while marked DISPATCHED_TO_FACTORY",
        )

    now = datetime.now(timezone.utc)
    event_role = select_role_for_action(user.roles, preferred=[Role.DISPATCH, Role.ADMIN])
    previous_status = job.current_status
    job.current_status = Status.PACKED_READY
    job.current_holder_role = STATUS_HOLDER_ROLE[Status.PACKED_READY]
    job.current_holder_user_id = user.id
    job.last_scan_at = now
    job.factory_id = None

    db.add(
        StatusEvent(
            job_id=job.id,
            from_status=previous_status,
            to_status=Status.PACKED_READY,
            scanned_by_user_id=user.id,
            scanned_by_role=event_role,
            timestamp=now,
            remarks=f"Removed from voucher {batch.batch_code}",
            override_reason="Removed from voucher",
        )
    )
    db.delete(batch_item)


def _sorted_batch_items(batch: Batch) -> list[BatchItem]:
    return sorted(
        batch.items,
        key=lambda item: item.added_at or datetime.min.replace(tzinfo=timezone.utc),
    )


def _build_manifest_workbook(batch: Batch):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    items = _sorted_batch_items(batch)
    total_weight = sum(float(item.job.approximate_weight or 0) for item in items if item.job)
    total_carat = sum(float(item.job.diamond_cent or 0) / 100 for item in items if item.job)
    total_amount = sum(float(item.job.purchase_value or 0) for item in items if item.job)

    wb = Workbook()
    summary = wb.active
    summary.title = "Voucher"
    summary.append(["Field", "Value"])
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)
    summary_rows = [
        ("Voucher Code", batch.batch_code),
        ("Status", batch.status.value if batch.status else None),
        ("Factory", batch.factory_name),
        ("Created At", batch.created_at.isoformat() if batch.created_at else None),
        ("Dispatch Date", batch.dispatch_date.isoformat() if batch.dispatch_date else None),
        ("Expected Return", batch.expected_return_date.isoformat() if batch.expected_return_date else None),
        ("Item Count", batch.item_count),
        ("Total Weight (g)", total_weight),
        ("Total Carat (ct)", total_carat),
        ("Total Value/Amount", total_amount),
    ]
    for row in summary_rows:
        summary.append(list(row))
    for row_idx in range(2, summary.max_row + 1):
        summary.cell(row=row_idx, column=1).font = Font(bold=True)
    for label in ("Total Weight (g)", "Total Carat (ct)", "Total Value/Amount"):
        for row_idx in range(2, summary.max_row + 1):
            if summary.cell(row=row_idx, column=1).value == label:
                summary.cell(row=row_idx, column=2).number_format = "#,##0.00"
                break
    summary.column_dimensions["A"].width = 22
    summary.column_dimensions["B"].width = 24

    ws = wb.create_sheet("Items")
    ws.append(MANIFEST_ITEM_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for batch_item in items:
        job = batch_item.job
        item_carat = float(job.diamond_cent or 0) / 100 if job.diamond_cent is not None else None
        ws.append([
            job.job_id,
            job.item_description,
            job.approximate_weight,
            item_carat,
            job.purchase_value,
            job.voucher_no,
            job.customer_name,
            job.customer_phone,
            job.style_number,
            job.card_weight,
            job.factory_name or batch.factory_name,
            job.current_status.value if job.current_status else None,
            job.work_narration,
            job.item_source.value if job.item_source else None,
            job.repair_type.value if job.repair_type else None,
            job.target_return_date.isoformat() if job.target_return_date else None,
            job.created_at.isoformat() if job.created_at else None,
            batch_item.added_at.isoformat() if batch_item.added_at else None,
        ])
        current_row = ws.max_row
        for column_idx in (3, 4, 5, 10):
            ws.cell(row=current_row, column=column_idx).number_format = "#,##0.00"

    ws.auto_filter.ref = f"A1:R{max(ws.max_row, 1)}"

    totals_start_row = ws.max_row + 2
    ws.cell(row=totals_start_row, column=1, value="Aggregate Totals").font = Font(bold=True)

    totals_rows = [
        ("Total Weight (g)", total_weight),
        ("Total Carat (ct)", total_carat),
        ("Total Value/Amount", total_amount),
    ]
    for offset, (label, value) in enumerate(totals_rows, start=1):
        label_cell = ws.cell(row=totals_start_row + offset, column=1, value=label)
        value_cell = ws.cell(row=totals_start_row + offset, column=2, value=value)
        label_cell.font = Font(bold=True)
        value_cell.font = Font(bold=True)
        value_cell.number_format = "#,##0.00"

    wb.active = wb.index(ws)
    return wb


@router.post("", response_model=BatchOut)
def create_batch(payload: BatchCreate, user=Depends(require_roles(Role.DISPATCH, Role.ADMIN)), db: Session = Depends(get_db)):
    now = datetime.now(timezone.utc)
    year = payload.year or now.year
    month = payload.month or now.month
    existing_codes = [
        row[0]
        for row in (
            db.query(Batch.batch_code)
            .filter(
                or_(
                    Batch.batch_code.like(f"VCH-{year}-{month:02d}-%"),
                    Batch.batch_code.like(f"BATCH-{year}-{month:02d}%"),
                )
            )
            .all()
        )
    ]
    batch_code = format_voucher_code(
        year,
        month,
        next_voucher_sequence(existing_codes, year=year, month=month),
    )
    factory_id = None
    if payload.factory_id:
        factory_id = _get_factory_by_uuid(db, payload.factory_id).id

    branch = _get_default_branch(db)
    batch = Batch(
        batch_code=batch_code,
        branch_id=branch.id,
        created_by=user.id,
        factory_id=factory_id,
        expected_return_date=payload.expected_return_date,
        status=BatchStatus.CREATED,
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)
    return batch


@router.get("", response_model=list[BatchOut])
def list_batches(db: Session = Depends(get_db), user=Depends(require_roles(Role.ADMIN, Role.DISPATCH, Role.FACTORY, Role.QC_STOCK))):
    return (
        db.query(Batch)
        .options(selectinload(Batch.factory))
        .order_by(Batch.created_at.desc())
        .limit(200)
        .all()
    )


@router.post("/{batch_id}/items", response_model=BatchOut)
def add_item(batch_id: str, payload: BatchAddItem, user=Depends(require_roles(Role.DISPATCH, Role.ADMIN)), db: Session = Depends(get_db)):
    batch = _get_batch(db, batch_id)
    job = _get_job_by_code(db, payload.job_id)
    if job.current_status != Status.PACKED_READY:
        raise HTTPException(status_code=400, detail="Only PACKED_READY items can be added")

    existing = db.query(BatchItem).filter(BatchItem.batch_id == batch.id, BatchItem.job_id == job.id).first()
    if existing:
        raise HTTPException(status_code=400, detail="Item already in voucher")

    db.add(BatchItem(batch_id=batch.id, job_id=job.id))
    batch.item_count += 1
    db.commit()
    db.refresh(batch)
    return batch


@router.post("/{batch_id}/dispatch", response_model=BatchOut)
def dispatch_batch(batch_id: str, payload: BatchDispatchRequest, user=Depends(require_roles(Role.DISPATCH, Role.ADMIN)), db: Session = Depends(get_db)):
    batch = _get_batch(db, batch_id)
    items = (
        db.query(BatchItem)
        .options(selectinload(BatchItem.job))
        .filter(BatchItem.batch_id == batch.id)
        .all()
    )
    if not items:
        raise HTTPException(status_code=400, detail="Voucher has no items")
    if payload.factory_id:
        factory = _get_factory_by_uuid(db, payload.factory_id)
        if batch.factory_id and batch.factory_id != factory.id:
            raise HTTPException(status_code=400, detail="Voucher factory does not match")
        batch.factory_id = factory.id
    if not batch.factory_id:
        raise HTTPException(status_code=400, detail="Factory id required before dispatch")

    jobs = [item.job for item in items]
    allowed_statuses = {
        Status.DISPATCHED_TO_FACTORY,
        Status.RECEIVED_AT_FACTORY,
        Status.RETURNED_FROM_FACTORY,
        Status.RECEIVED_AT_SHOP,
        Status.ADDED_TO_STOCK,
        Status.HANDED_TO_DELIVERY,
        Status.DELIVERED_TO_CUSTOMER,
    }
    for job in jobs:
        if job.current_status not in allowed_statuses:
            raise HTTPException(status_code=400, detail=f"Job {job.job_id} is not dispatched yet")

    batch.status = BatchStatus.DISPATCHED
    if payload.dispatch_date:
        batch.dispatch_date = payload.dispatch_date
    elif not batch.dispatch_date:
        batch.dispatch_date = datetime.now(timezone.utc)
    if payload.expected_return_date:
        batch.expected_return_date = payload.expected_return_date
        for job in jobs:
            if job.target_return_date is None:
                job.target_return_date = payload.expected_return_date
    db.commit()
    db.refresh(batch)
    return batch


@router.get("/{batch_id}", response_model=BatchDetail)
def get_batch(batch_id: str, db: Session = Depends(get_db), user=Depends(require_roles(Role.ADMIN, Role.DISPATCH, Role.FACTORY, Role.QC_STOCK))):
    batch = _get_batch(db, batch_id, with_items=True, with_factory=True)
    items = [item.job for item in batch.items]
    return BatchDetail(
        **BatchOut.model_validate(batch).model_dump(),
        items=[JobOut.model_validate(job) for job in items],
    )


@router.delete("/{batch_id}/items/{job_id}", response_model=BatchOut)
def remove_item(
    batch_id: str,
    job_id: str,
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.ADMIN, Role.DISPATCH)),
):
    batch = _get_batch(db, batch_id)
    if batch.status == BatchStatus.CLOSED:
        raise HTTPException(status_code=400, detail="Voucher is closed")

    batch_item = (
        db.query(BatchItem)
        .options(selectinload(BatchItem.job))
        .join(ItemJob, BatchItem.job_id == ItemJob.id)
        .filter(BatchItem.batch_id == batch.id, ItemJob.job_id == job_id)
        .first()
    )
    if not batch_item:
        raise HTTPException(status_code=404, detail="Item not found in voucher")

    _remove_batch_item(batch, batch_item, user)
    _sync_batch_item_count(db, batch)
    db.commit()
    db.refresh(batch)
    return batch


@router.delete("/{batch_id}/items", response_model=BatchOut)
def clear_batch_items(
    batch_id: str,
    db: Session = Depends(get_db),
    user=Depends(require_roles(Role.ADMIN, Role.DISPATCH)),
):
    batch = _get_batch(db, batch_id)
    if batch.status == BatchStatus.CLOSED:
        raise HTTPException(status_code=400, detail="Voucher is closed")

    items = (
        db.query(BatchItem)
        .options(selectinload(BatchItem.job))
        .filter(BatchItem.batch_id == batch.id)
        .all()
    )
    for batch_item in items:
        _remove_batch_item(batch, batch_item, user)

    _sync_batch_item_count(db, batch)
    db.commit()
    db.refresh(batch)
    return batch


@router.get("/{batch_id}/manifest.pdf")
def get_manifest(batch_id: str, db: Session = Depends(get_db), user=Depends(require_roles(Role.ADMIN, Role.DISPATCH))):
    batch = _get_batch(db, batch_id, with_items=True, with_factory=True)
    jobs = [item.job for item in batch.items]
    pdf_bytes = generate_manifest_pdf(batch, jobs)
    return StreamingResponse(iter([pdf_bytes]), media_type="application/pdf")


@router.get("/{batch_id}/manifest.xlsx")
def get_manifest_xlsx(batch_id: str, db: Session = Depends(get_db), user=Depends(require_roles(Role.ADMIN, Role.DISPATCH))):
    batch = _get_batch(db, batch_id, with_items=True, with_factory=True)
    wb = _build_manifest_workbook(batch)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'{batch.batch_code.lower().replace(" ", "-")}-manifest.xlsx'
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/{batch_id}/close", response_model=BatchOut)
def close_batch(batch_id: str, db: Session = Depends(get_db), user=Depends(require_roles(Role.ADMIN, Role.DISPATCH))):
    batch = _get_batch(db, batch_id)
    jobs = [item.job for item in batch.items]
    allowed_statuses = {
        Status.RECEIVED_AT_SHOP,
        Status.ADDED_TO_STOCK,
        Status.HANDED_TO_DELIVERY,
        Status.DELIVERED_TO_CUSTOMER,
    }
    for job in jobs:
        if job.current_status not in allowed_statuses:
            raise HTTPException(status_code=400, detail="Not all items have returned")

    batch.status = BatchStatus.CLOSED
    db.commit()
    db.refresh(batch)
    return batch
